# -*- coding: utf-8 -*-
"""
Created on Wed Jun  1 09:06:51 2022

@author: vladimirov.aa

requirements.txt -
numpy==1.21.5
openpyxl==3.0.9
pandas==1.4.2

"""
#%%
'''
IMPORTS
'''
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog

import datetime as dt
#%%
'''
FRONTEND
'''
class MainApp(tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        
        self.geometry('600x230')
        self.title("Предобработка данных для бизнес-кейса")
        self.resizable(False, False)
        
        container = tk.Frame(self)
        container.pack(fill='both', expand=True)
       
        self.frames={}
        for frame_class in (Bottombar, Checkboxes, Radiobuttons, ):
           frame = frame_class(container, self)
           self.frames[frame_class] = frame
           
           frame.pack(side='bottom')
           if frame_class in (Bottombar, Checkboxes): ttk.Separator(container, orient='horizontal').pack(side='bottom',fill='x') 
            
    def switch_widgets(self, state, widgets):
        states = {'normal', 'disabled'}
        if state not in states:
            raise ValueError("switch_widgets() argument state should be 'disabled' or 'normal")
        for widget in widgets:
            widget.configure(state=state)
            widget.update()

    def get_frame(self, frame_class):
       return self.frames[frame_class]
   
    def openfile(self):
        self.uploaded_xlsx = filedialog.askopenfilename(initialdir = "./Входные данные",title = "Select file",
                                         filetypes = (("Excel","*.xlsx"),("all files","*.*")))
    
    def calculate(self):
        Calculate(self, self.frames)
       
class Bottombar(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent, borderwidth=5)
        self.controller = controller
        
        self.nmx_start = tk.Label(self, text= "NumEx Сутки 0")
        self.nmx_entry_month = tk.Spinbox(self, from_=1, to_=12, width=2)
        self.nmx_entry_year = tk.Spinbox(self, from_=1900, to_=2200, width=4) 
        self.nmx_div1 = tk.Label(self, text="мес")
        self.nmx_div2 = tk.Label(self, text="год")        
        
        self.widgets2switch = (self.nmx_start, self.nmx_entry_month, self.nmx_entry_year,
                               self.nmx_div1, self.nmx_div2)
        self.controller.switch_widgets('disabled', self.widgets2switch)
        
        self.nmx_start.grid(row=0, column=0, columnspan=4)
        self.nmx_entry_month.grid(row = 1, column = 1)
        self.nmx_entry_year.grid(row = 1, column = 3)
        self.nmx_div1.grid(row=1, column=0)
        self.nmx_div2.grid(row=1, column=2)    
        
      
        tk.Label(self,text= "Номер сектора:").grid(row = 0, column = 5, sticky='w')
        tk.Label(self,text= "Старт расчета:").grid(row = 1, column = 5, sticky='w')
        
        self.sector_entry = tk.Entry(self, width=20)  
        #self.sector_entry.insert(0, "1") # default value
        self.month_entry = tk.Spinbox(self, from_=1, to_=12, width=2)
        self.year_entry = tk.Spinbox(self, from_=dt.datetime.now().year, 
                                to_=dt.datetime.now().year+100, width=4)        
        
        self.sector_entry.grid(row=0, column=6, columnspan=4)
        self.month_entry.grid(row=1, column=7)
        self.year_entry.grid(row=1, column=9)
    
        tk.Label(self, text="мес").grid(row=1, column=6)
        tk.Label(self, text="год").grid(row=1, column=8)
    
        tk.Label(self, text="", width=3).grid(row=0, column=10, rowspan=2)
        tk.Label(self, text="", width=3).grid(row=0, column=4, rowspan=2)
    
    
        tk.Button(self, width=20, text='Открыть файл (.xlsx)',
                  command=self.controller.openfile).grid(row=0, column=11, sticky='E')
        tk.Button(self, width=20, text='Расчет (Enter)', 
                  command=self.controller.calculate).grid(row=1, column=11, sticky='E')
        
        self.controller.bind('<Return>', lambda event: self.controller.calculate())
      
class Checkboxes(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent, borderwidth=5)
        self.controller = controller

        tk.Label(self, text="Пересчет профилей\nна данные МЭР:",)\
            .grid(row=0, column=0, rowspan=3)
          
        txts = ("База,\nдебит нефти",  "База,\nдебит жидкости", "База,\nприемистость",  
                "ОБД,\nдебит нефти",   "ОБД,\nдебит жидкости",  "ОБД,\nприемистость")

        rows_grid, cols_grid = (2, 3)

        counter=0
        self.var_dict = {}
        for i in range(rows_grid):
            for j in range(cols_grid):
                self.var_dict[i,j] = tk.IntVar(value=1) #chbox is checked by default
                tk.Checkbutton(self, text=txts[counter], variable=self.var_dict[i,j], justify=tk.LEFT,
                               onvalue=1, offvalue=0, padx=15, pady=10).grid(row=i, column=j+1)
                counter += 1

class Radiobuttons(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent, borderwidth=5)
        self.controller = controller
 
        tk.Label(self, text= "Формат расчета:")\
            .grid(row = 1 , column = 0)

        self.dtfrmt = tk.StringVar(self, "год") # значение в радиокнопке по умолчанию
        tk.Radiobutton(self, value='месяц', text="по месяцам", variable=self.dtfrmt).grid(row=1, column=1)
        tk.Radiobutton(self, value='год', text="по годам", variable=self.dtfrmt).grid(row=1, column=2)
                     
        tk.Label(self, text='    ').grid(row=1, column=3)
            
        tk.Label(self, text='Формат входных данных:')\
            .grid(row = 1, column = 4)
        self.widgets2switch = self.controller.get_frame(Bottombar).widgets2switch

        self.profile_frmt = tk.StringVar(self, 'tnav') # значение в радиокнопке по умолчанию
        
        tk.Radiobutton(self, value='tnav', text="tNav", variable=self.profile_frmt,
                       command=lambda: self.controller.switch_widgets('disabled', self.widgets2switch)
                       ).grid(row=1, column=5) #disableEntry
        tk.Radiobutton(self, value='numex', text="NumEx", variable=self.profile_frmt,
                       command=lambda: self.controller.switch_widgets('normal', self.widgets2switch)
                       ).grid(row=1, column=6) #enableEntry
        
#%%
'''
IMPORTS
'''
import pandas as pd       
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

import warnings
def warning_on_one_line(message, category, filename, lineno, file=None, line=None):
    return ' %s:%s: %s:%s\n\n' % (filename, lineno, category.__name__, message)
warnings.formatwarning = warning_on_one_line
warnings.simplefilter(action='ignore', category=pd.errors.PerformanceWarning)

#%%
'''
BACKEND
'''
class Calculate():
    def __init__(self, mainapp, frames):
        self.mainapp = mainapp
        self.frames = frames
        print('Uploading .xlsx file...')
        self.upload()
        self.checkformats()
        print("... headers' formats are OK.\n")
        print('Calculation started...')
        self.prepare()
        self.write()
        self.destroy() #global nmx variables
        print('(имена скважин преобразованы в соответствии со столбцом "ИМЯ_OIS" из шахматки)')
        print('Calculated!')
                
    def upload(self):     
        # введеные пользователем данные
        global sector, month, year, nmx_month, nmx_year
        try:
            sector = int(self.frames[Bottombar].sector_entry.get())
            month = int(self.frames[Bottombar].month_entry.get())
            year = int(self.frames[Bottombar].year_entry.get())
            
            if self.frames[Radiobuttons].profile_frmt.get() == 'numex':
                nmx_month = int(self.frames[Bottombar].nmx_entry_month.get())
                nmx_year = int(self.frames[Bottombar].nmx_entry_year.get())
        except:
            raise ValueError('Entered value is not an integer')   

        key_map = {(0, 0):('База', 'деб_неф'),  (0, 1):('База', 'деб_жидк'),  (0, 2):('База', 'приемист'),  
                   (1, 0):('ОБД',  'деб_неф'),  (1, 1):('ОБД',  'деб_жидк'),  (1, 2):('ОБД',  'приемист'), }

        global chbox_dict
        chbox_dict = {key_map[key]: value.get() for (key,value) in self.frames[Checkboxes].var_dict.items()}
           
        global exportfrmt, profilefrmt
        exportfrmt = self.frames[Radiobuttons].dtfrmt.get() #выбираем расчет таблиц 'месяц' или 'год'
        profilefrmt = self.frames[Radiobuttons].profile_frmt.get()

        # данные из Excel файла
        try:
            excel = pd.ExcelFile(self.mainapp.uploaded_xlsx)
        except:
            raise FileNotFoundError('.xlsx was not found')
        
        # данные для справочников
        if profilefrmt == 'tnav': # матрицы дренирования
            self.baza_mtrx = excel.parse('Матрица_дренирования_база', header=1)
            self.obd_mtrx = excel.parse('Матрица_дренирования_обд', header=1)
            
        elif profilefrmt == 'numex': # координаты скважин
            self.coords = excel.parse('скважины')
            self.coords.rename(columns={self.coords.columns[0]:'скв'}, inplace=True)
            self.coords['скв'] = self.coords['скв'].astype(str)
        
        # данные для профилей
        self.baza_profile = excel.parse('База_профиль')
        self.obd_profile = excel.parse('ОБД_профиль')
                
        self.MER, self.chequer = excel.parse('МЭР'), excel.parse('шахматка')
        
        self.chequer['Номер сектор'] = self.chequer['Номер сектор'].astype(int)
        if not sector in self.chequer['Номер сектор'].unique():
            raise ValueError('Заданный номер сектора отсутсвует во входном файле')
     
        
    def checkformats(self):
        if profilefrmt == 'tnav':
            if not all(item in ['Скважина', 'Флюид (пл. усл.), пласт.м3', 'Нефть, ст.м3', 'Скважина.1','Флюид (пл. усл.), пласт.м3.1', 'Нефть, ст.м3.1'] for item in self.obd_mtrx.columns.union(self.baza_mtrx.columns)):
                raise ValueError('Ошибка в заголовках матриц дренирования')
            if not all(item in ['Объект', 'Шаг', 'Дата', 'Дней', 'Массовый дебит нефти, т/сут', 'Массовый дебит воды, т/сут', 'Приёмистость воды, ст.м3/сут','Время работы на текущем временном шаге, сут.'] for item in self.obd_profile.columns.union(self.baza_profile.columns)):
                raise ValueError('Ошибка в заголовках профилей Базы или ОБД')

        if profilefrmt == 'numex':
            if not all(item in self.coords.columns for item in ['скв', 'доб(1)/наг(2)', 'X1', 'Y1']):
                raise ValueError('Ошибка в заголовках листа с данными скважин')
            if not all(item in self.obd_profile.columns.union(self.baza_profile.columns) for item in ['Скв', 'Сутки', 'Деб_жид_(т/сут)', 'Деб_неф_(т/сут)','Деб_газа_(тыс.м3/сут)', 'Обв_(м3/м3)', 'Нак_жид_(т)', 'Нак_неф_(т)','Нак_газ_(тыс.м3)', 'Зак_вод_(м3/сут)', 'Зак_газ_(тыс.м3/сут)','Нак_закв_(м3)', 'Нак_закг_(тыс.м3)', 'Заб_дав_(бар)', 'СДФ_доб','СДФ_наг', 'Куст']):
                raise ValueError('Ошибка в заголовках профилей Базы или ОБД')
        
        if not all(item in ['Скв. модель', 'Qн, т/сут', 'Qв, т/сут', 'Qж, т/сут', 'Приемистость, м3/сут'] for item in self.MER.columns):
            raise ValueError('Ошибка в заголовках МЭР')
        if not all(item in ['FIELD', 'ИМЯ_OIS', 'ИМЯ_NGT', 'ГДМ', 'Номер сектор', 'КУСТ', 'СТВОЛ_В_РАБОТЕ', 'ГФ', 'Группа'] for item in self.chequer.columns):
            raise ValueError('Ошибка в заголовках шахматки')

       
    def prepare(self):        
        # профиля       
        self.bazpr = self.get_profile(self.baza_profile)
        self.obdpr = self.get_profile(self.obd_profile)
    
        # МЭР
        self.MER.columns = ['объект', 'деб_неф', 'деб_вод', 'деб_жидк', 'приемист']
        self.MER['объект'] = self.MER['объект'].astype(str)
        self.MER.set_index('объект', inplace=True)
    
        # шахматка 
        self.chequer = self.chequer[(self.chequer['СТВОЛ_В_РАБОТЕ'] == 1) &
                              (self.chequer['Номер сектор'] == sector)]
        self.chequer[['ИМЯ_OIS', 'ИМЯ_NGT', 'ГДМ', 'КУСТ']] \
            = self.chequer[['ИМЯ_OIS', 'ИМЯ_NGT', 'ГДМ', 'КУСТ']].astype(str)
            
        # проверка профилей
        if not set(self.bazpr['объект']).issubset(set(self.chequer['ГДМ'])):
            warnings.warn('Профиль БАЗА - список скважин из профиля не является подмножеством списка скважин из столбца "ГДМ" шахматки. Отсутствующие в шахматке скважины будут отсечены.')
            self.bazpr = self.bazpr[self.bazpr['объект'].isin(self.chequer['ГДМ'])]
        if not set(self.obdpr['объект']).issubset(set(self.chequer['ГДМ'])):
            warnings.warn('Профиль ОБД - список скважин из профиля не является подмножеством списка скважин из столбца "ГДМ" шахматки. Отсутствующие в шахматке скважины будут отсечены.')
            self.obdpr = self.obdpr[self.obdpr['объект'].isin(self.chequer['ГДМ'])]

        self.bazpr.name = 'База' # будет использоваться в шапке выходного файла
        self.obdpr.name = 'ОБД' # будет использоваться в шапке выходного файла
        

    def get_profile(self, profile):
        '''
        подготовка и обработка профиля
        '''
        profile = profile.copy()
            
        profile = profile[profile.filter(regex='^(?!Unnamed)').columns] #убрать столбцы с NaN
        profile.dropna(inplace=True)

        # проверка источника данных
        # названия столбцов приводятся к формату Инструмента (ФЭМ)
        if profilefrmt == 'tnav':
            profile.rename(columns = 
                           {'Объект':'объект', 'Дата':'дата', 'Массовый дебит нефти, т/сут':'деб_неф',
                           'Массовый дебит воды, т/сут':'деб_вод', 'Приёмистость воды, ст.м3/сут':'приемист',
                           'Время работы на текущем временном шаге, сут.':'время'}, inplace=True)

            # удалим строчки с бездействующими скважинами
            profile = profile[profile['время'].ne(0)]
        
            # удалим строчки, в ячейках которых есть хотя бы одно значение,
            # полностью состоящее из буквенных символов
            profile = profile[~profile[[ 'дата', 'деб_неф', 'деб_вод', 'приемист', 'время']]\
                              .applymap(lambda x: str(x).isalpha()).any(1)]#.reset_index()
                
            # пересчитаем значения профиля
            profile['деб_жидк'] = profile['деб_вод'] + profile['деб_неф'] # т/сут
            profile['дбч_неф'] = profile['деб_неф'] * profile['время'] / 1000 # тыс. тонн
            profile['дбч_жидк'] = profile['деб_жидк'] * profile['время'] / 1000 # тыс. тонн
            
            GF = self.chequer.set_index('ГДМ')['ГФ'].rename('ГФ')
            profile = profile.join(GF, on='объект')
            profile['дбч_газ'] = profile['дбч_неф'] * profile['ГФ'] / 1000
            profile['закачка'] = profile['приемист'] * profile['время'] / 1000 # тыс. м3
            
        elif profilefrmt == 'numex':
            profile.rename(columns = {'Скв':'объект', 'Деб_неф_(т/сут)':'деб_неф', 'Деб_жид_(т/сут)':'деб_жидк', 
                                      'Деб_газа_(тыс.м3/сут)':'деб_газ', 'Нак_газ_(тыс.м3)':'дбч_газ',
                                      'Зак_вод_(м3/сут)': 'приемист', 'Нак_неф_(т)':'дбч_неф', 'Нак_жид_(т)':'дбч_жидк',
                                      'Нак_закв_(м3)':'закачка'}, inplace=True)
           
            # удалим строчки с неработающими в этом месяце скважинами
            profile = profile[profile[['деб_неф', 'деб_жидк', 'деб_газ', 'приемист']].ne(0).any(1)]
            
            startdate = dt.date(nmx_year, nmx_month, 1)
            profile['дата'] = startdate + (profile['Сутки']/30).apply(pd.offsets.MonthBegin)

        # TODO            
        # подвинем даты на месяц назад
        profile['дата'] = profile['дата'] - pd.DateOffset(months=1)

        profile = profile[['объект', 'дата', 'деб_неф', 'деб_жидк', 'приемист', 
                           'дбч_неф', 'дбч_жидк', 'дбч_газ', 'закачка']]
        
        profile['объект'] = profile['объект'].astype(str)

        profile.rename(columns = {'дата':'месяц'}, inplace=True)
        profile = profile[profile['месяц'].map(lambda x: x.month) >= month]
        profile['год'] = profile['месяц'].map(lambda x: x.year)
        profile = profile[profile['год'] >= year]

        if (profilefrmt == 'numex') & profile.empty:
            raise ValueError('Загруженный профиль не совпадает по датам. Попробуйте изменить Сутки 0.')
       
        # добавим столбец для КРС заглушки (временное решение)
        profile['КРС'] = None
        
        global header_width
        header_width = profile[exportfrmt].nunique()
        
        return profile
        
      
    def write(self):
        out_path = 'Расчет_экономики_'+ exportfrmt +'.xlsx'
        
        self.writer = pd.ExcelWriter(out_path, engine='xlsxwriter',
                                datetime_format='mmm.yy') #формат даты для Добыча_мес ('mmm.yy'))
        # Справочники (2 листа Excel)
        GuideHandler(self).write(['база','обд'])
        
        self.sheetname_deb = {'год':'Добыча', 'месяц':'Добыча_мес'}
        # Добыча (1 лист Excel)
        ProfileHandler(self).write([self.obdpr, self.bazpr], ['база','обд'], ['деб_неф', 'деб_жидк'], 
                                   self.sheetname_deb[exportfrmt], PNG=True)
                
        self.sheetname_nag = {'год':'Закачка', 'месяц':'Закачка_мес'}
        # Закачка (1 лист Excel)
        ProfileHandler(self).write([self.obdpr, self.bazpr], ['база','обд'], ['приемист'], 
                                   self.sheetname_nag[exportfrmt])

        if exportfrmt == 'месяц': 
            # КРС (1 лист, только для формата 'по месяцам')
            ProfileHandler(self).write_KRS([self.obdpr, self.bazpr], ['база','обд'], ['КРС'],
                                          sheet_name='КРС')
        
        # Листы с дебитами и приемистостью (6 листов, опционально)
        sheet_name_prm = {'деб_неф':'дебиты нефть', 'деб_жидк':'дебиты жидкость', 'приемист':'приемистость'}
        for profile in [self.obdpr, self.bazpr]:
            for prm in ['деб_неф', 'деб_жидк', 'приемист']:        
                ProfileHandler(self).write(profile, profile.name, prm, 
                                           sheet_name=profile.name+' '+sheet_name_prm[prm],
                                           single_pivtab=True)

        # save file
        self.writer.save()
        
        # open saved .xlsx in openpyxl
        self.wb = openpyxl.load_workbook(filename = 'Расчет_экономики_'+ exportfrmt +'.xlsx')

        self.fix_multiindex()
        self.format_cells()
        
        # save formated file
        self.wb.save(filename = 'Расчет_экономики_'+ exportfrmt +'.xlsx')                
    
        
    def fix_multiindex(self):
        '''
        костыль метода to_excel() при записи  Multi-Index шапок - 
        компенсируем недостающий функционал с помощью openpyxl :(
        '''
        def delete_col_with_merged_ranges(sheet, idx): #решаем проблему с сохранением формата шапки
            sheet.delete_cols(idx)
            for mcr in sheet.merged_cells:
                if idx < mcr.min_col:
                    mcr.shift(col_shift=-1)
                elif idx <= mcr.max_col:
                    mcr.shrink(right=1)
                    
        delete_col_with_merged_ranges(self.wb[self.sheetname_deb[exportfrmt]], 1)
        self.wb[self.sheetname_deb[exportfrmt]].delete_rows(7)
        
        delete_col_with_merged_ranges(self.wb[self.sheetname_nag[exportfrmt]], 1)
        self.wb[self.sheetname_nag[exportfrmt]].delete_rows(7)
        
        if exportfrmt == 'месяц':
            self.wb['КРС'].delete_rows(7)
            delete_col_with_merged_ranges(self.wb['КРС'], 1)

        
    def format_cells(self):
        '''
        Format cells using openpyxl
        '''
        for wsi in [self.wb[self.sheetname_deb[exportfrmt]], self.wb[self.sheetname_nag[exportfrmt]]]:
            ws = wsi
            pane_width = 6 if ('Добыча' in ws.title) & (profilefrmt=='tnav') else 5
            
            for rows in  ws.iter_rows(min_row=4, max_row=6, min_col=pane_width+1):
                snippet=[]
                for i in range(0, int(len(rows)/header_width), 2):
                     snippet += rows[i*header_width : (i+1)*header_width]
                for cell in snippet:
                    cell.fill = PatternFill(fgColor='DCDCDC', fill_type='solid')
            
            for idx, col in enumerate(ws.columns, pane_width+1):
                ws.column_dimensions[get_column_letter(idx)].auto_size = True
                ws.column_dimensions[get_column_letter(idx)].number_format = '0.00E+00'
                
                    
    def destroy(self):
        if exportfrmt=='numex':
            print(nmx_month,  nmx_year) 
            del globals()['nmx_month', 'nmx_year'] # automatically destroy values of previous iteration
        
        
class GuideHandler:        
    def __init__(self, parent):
        self.parent = parent
        
    def write(self, guide_names): #напишем dummy для избегания конфликта имен
        for name in guide_names:
            sheet = self.get_guide(name)
            sheet.to_excel(self.parent.writer, sheet_name='Справочник ДОБ-НАГ '+ name, startrow=1, index=False)
        
    def get_guide(self, name):
        '''
        если profilefrmt = 'tnav':   
            Подсчитывает справочник с коэффициентами влияния на основе матриц дренирования
        иначе если profilefrmt = 'numex':
            Подсчитывает справочник с коэффициентами влияния на основе координат скважин
        '''
        if profilefrmt == 'tnav':
            guide = self.parent.baza_mtrx.copy() if name=='база' else self.parent.obd_mtrx.copy()# if name=='обд')
            
            guide.columns = ['наг_скв', 'наг_флюид', 'наг_нефть', #[ , м3, м3,
                             'доб_скв', 'доб_флюид', 'доб_нефть'] #  , м3, м3,]
            
            guide.fillna(method='ffill', inplace = True)
            guide[['наг_скв', 'доб_скв']] = guide[['наг_скв', 'доб_скв']].astype(str)
            
            guide = guide[~guide['наг_скв'].isin(['Резервуар'])]
            ireservoir = guide[guide['доб_скв'].isin(['Резервуар'])].index
                
            guide['разница_флюидов'] = guide.loc[ireservoir]['наг_флюид']\
                                        - guide.loc[ireservoir]['доб_флюид']
            guide.fillna(method='ffill', inplace = True)                           
            
            guide['КЭФ'] = guide['доб_флюид'] / guide['разница_флюидов']
            
            guide = guide[guide['доб_скв'].isin(['Резервуар']) == False]
            
            guide = guide[['доб_скв', 'наг_скв', 'КЭФ']]
            guide.columns = ['ДОБ', 'НАГ', 'КЭФ']
        

        
        elif profilefrmt == 'numex':
            profile = self.parent.bazpr.copy() if name=='база' else self.parent.obdpr.copy()
            coords = self.parent.coords.copy()
            R = 1000 # радиус влияния наг скважин

            well_lst = profile['объект'].unique().tolist()
            coords = coords[coords['скв'].isin(well_lst)]
            nag = coords[coords['доб(1)/наг(2)']==2][['скв', 'X1', 'Y1']]
            dob = coords[coords['доб(1)/наг(2)']==1][['скв', 'X1', 'Y1']]
            
            rads = nag.merge(dob, how='cross', suffixes=('_nag','_dob'))
            rads['r'] = np.sqrt(((rads.X1_nag-rads.X1_dob)**2 + (rads.Y1_nag-rads.Y1_dob)**2))
            
            influenced = rads[rads.r <= R].copy()
            influenced['1/r**2'] = 1/influenced.r**2
            
            SUM = influenced.groupby(['скв_nag'])['1/r**2'].sum().rename('SUM')
            
            influenced = influenced.merge(SUM, left_on='скв_nag', right_index=True)
            influenced['КЭФ'] = influenced['1/r**2']/influenced['SUM']

            guide = influenced[['скв_dob','скв_nag','КЭФ']].copy()
            guide.columns = ['ДОБ', 'НАГ', 'КЭФ']

        else:
            raise ValueError("Profile format should be 'tnav' or 'numex'")
        
        msg = "матриц дренирования" if profilefrmt =='tnav' else "профилей добычи/закачки"
        
        if not set(pd.concat([guide['ДОБ'], guide['НАГ']])).issubset(set(self.parent.chequer['ГДМ'])):
            warnings.warn(f'Справочник {name.upper()} - список скважин из {msg} не является подмножеством списка скважин столбца "ГДМ" шахматки. Отсутствующие в шахматке скважины будут отсечены.')
            guide = guide[guide['ДОБ'].isin(self.parent.chequer['ГДМ'])]
            guide = guide[guide['НАГ'].isin(self.parent.chequer['ГДМ'])]
        
        #if not guide['ДОБ'].isin(self.parent.chequer['ГДМ']).all():
        #    warnings.warn(f'Справочник {name.upper()} - список ДОБ скважин из {msg} не является подмножеством списка скважин столбца "ГДМ" шахматки. Отсутствующие в шахматке скважины будут отсечены.')
         #   guide = guide[guide['ДОБ'].isin(self.parent.chequer['ГДМ'])]
        #if not guide['НАГ'].isin(self.parent.chequer['ГДМ']).all():
         #   warnings.warn(f'Справочник {name.upper()} - список НАГ скважин из {msg} не является подмножеством списка скважин столбца "ГДМ" шахматки. Отсутствующие в шахматке скважины будут отсечены.')
          #  guide = guide[guide['НАГ'].isin(self.parent.chequer['ГДМ'])]
            
        # присвоим ИМЯ_OIS для номера скважины
        guide['ДОБ'] = self.parent.chequer.set_index('ГДМ').loc[guide['ДОБ']]['ИМЯ_OIS'].to_list()
        guide['НАГ'] = self.parent.chequer.set_index('ГДМ').loc[guide['НАГ']]['ИМЯ_OIS'].to_list()

        return guide

        
class ProfileHandler:
    def __init__(self, parent):
        self.parent = parent
        
    def write(self, profiles, pivtab_names, params, sheet_name:str, 
              PNG = False, single_pivtab=False):
        freeze_panes = (6,6) if (PNG==True)&(profilefrmt=='tnav') else (6,5)
        
        if single_pivtab==False:
            sheet = self.get_multiple_pivtabs_sheet(profiles, pivtab_names, params, PNG)
        elif single_pivtab==True:
            sheet = self.get_single_pivtab_sheet(profiles, pivtab_names, params, PNG)
        
        sheet.to_excel(self.parent.writer, sheet_name=sheet_name, startrow=3 if single_pivtab==False else 5, 
                       index=True, freeze_panes=freeze_panes if single_pivtab==False else None)
        
        
    def get_multiple_pivtabs_sheet(self, profiles, pivtab_names, params, PNG):
        '''
        profiles - [obdpr, bazpr] 
        prm in params - ['деб_неф','деб_жидк'],['приемист']
        PNG True or False - добавить столбец с ГФ
        '''
        # горизонтальная шапка сводной таблицы с привязкой к объектам
        pivtab_pane = self.get_pivtab_pane(profiles, PNG)
        # 2 и 3 уровень шапки
        pivtab_pane.columns = pd.MultiIndex.from_product([[''], [''], pivtab_pane.columns])
        
        
        dct_pvt = {} # словарь dictionary (d) с датафреймами pivtab (pt), построенных на профиле
        dct_prfls = {} # словарь (d) с датафреймами pivtab (pvt), построенных на dct_pvt
        
        #deb_dict = {'дбч_неф':'деб_неф', 'дбч_жидк':'деб_жидк', 
         #           'закачка':'приемист'}  # словарь соответсвия кумулят. дебитов и дебитов
        deb_dict = {'деб_неф':'дбч_неф', 'деб_жидк':'дбч_жидк', 
                    'приемист':'закачка'}  # словарь соответсвия кумулят. дебитов и дебитов
        
        header_dict = {'дбч_неф':'Qнефть ','дбч_жидк':'Qжидк ', 
                       'закачка':'Qжидк '} # словарь с шапкой дат для exportfrmt='год'
    
        for profile in profiles:          
            for prm in params:
                
                dct_pvt[prm] = profile.pivot_table(values=deb_dict[prm],
                           index='объект',columns=exportfrmt, aggfunc='sum')
    
                #if chbox_dict[profile.code, prm_dict[prm]] == 1:
                if chbox_dict[profile.name, prm] == 1:
                    # пересчитываем конечные значения на коэффициенты МЭР,
                    # высчитанные с использованием интересующего профиля
                    prm_corr = profile[['объект', prm, deb_dict[prm] ]].set_index('объект')
    
                    prm_corr = prm_corr[prm_corr[prm].ne(0)]\
                        .groupby('объект').first() #отсекаем нулевые значения
                        #оставляем только первые значения из группы
    
                    # подгружаем значения параметра из МЭР
                    MERval = self.get_MER_values(prm)                          
                    # считаем коэффициенты
                    coefs = MERval.div(prm_corr[[prm]]).fillna(1)
                    #пересчитываем сводную таблицу
                    dct_pvt[prm] = pd.merge(how='left', left=dct_pvt[prm], right=coefs, 
                                  left_index=True, right_index=True)
    
                    dct_pvt[prm] = dct_pvt[prm].apply(axis=1,
                                   func=lambda row: row.iloc[:-1]*row.loc[prm])
    
                # сразу поменяем формат заголовка внутри цикла
                if exportfrmt == 'год':
                    dct_pvt[prm].columns = dct_pvt[prm].columns.astype(str)\
                        .str.replace('\A', header_dict[deb_dict[prm]], regex=True) #привести строки к необх формату
                elif exportfrmt == 'месяц':
                    pass
    
                # добавляем 2 и 3 уровень шапки
                dct_pvt[prm].columns = pd.MultiIndex.from_product([[profile.name+' '+deb_dict[prm]],
                                                                 [''], dct_pvt[prm].columns])
            
            dct_prfls[profile.name] = pd.concat(dct_pvt.values(), axis=1)
        
        pivtab = pd.concat(dct_prfls.values(), axis=1)
        
        if PNG == False:
            pass
        else: # добавляем табличку с ПНГ
            if profilefrmt == 'numex':
                
                dct_PNG = {}
                for profile in profiles: 
                    
                    #рассчитываем ПНГ
                    dct_PNG[profile.name] = profile.pivot_table(values='дбч_газ', index='объект',
                                                                columns=exportfrmt, aggfunc='sum')
                    # добавляем 2 и 3 уровень шапки
                    dct_PNG[profile.name].columns = \
                        pd.MultiIndex.from_product([[profile.name+' ПНГ'], [''], 
                                                    dct_PNG[profile.name].columns])                                
                
                PNGtab = pd.concat(dct_PNG.values(), axis=1)
                
                pivtab = pd.merge(left=pivtab, right=PNGtab, how='left', 
                                  left_index=True, right_index=True, sort=True)
                
                                
            elif profilefrmt == 'tnav':
                dct_PNG = {}
                for profile in profiles: 
                    
                    dct_PNG[profile.name] = pivtab.loc[:,(profile.name+' дбч_неф', '')].copy()
                    GF = pivtab_pane[('', '', 'ГФ')].copy()
                    
                    #рассчитываем ПНГ
                    dct_PNG[profile.name] = dct_PNG[profile.name].mul(GF, axis=0)/1000 
                    
                    # добавляем 2 и 3 уровень шапки
                    dct_PNG[profile.name].columns = \
                        pd.MultiIndex.from_product([[profile.name+' ПНГ'], [''], 
                                                    dct_PNG[profile.name].columns])                                

                PNGtab = pd.concat(dct_PNG.values(), axis=1)
                
                pivtab = pd.merge(left=pivtab, right=PNGtab, how='left', 
                                  left_index=True, right_index=True, sort=True)

    
        pivtab = pivtab.fillna(0)
        pivtab = pivtab.loc[~(pivtab==0).all(axis=1)]
           
        # пересечем скважины из получившейся таблицы со скважинами из шахматки
        pivtab = pd.merge(left=pivtab_pane, right=pivtab, how='inner', 
                          left_index=True, right_index=True, sort=True)\
        
        # присвоим ИМЯ_OIS для номера скважины
        pivtab.loc[:,('','',['объект'])] = \
            self.parent.chequer.set_index('ГДМ').loc[pivtab.index]['ИМЯ_OIS']
        
        pivtab.rename(columns = 
                      {'группа':'Группа', 'версия':'Версия профиля', 'сектор':'№ сектора',
                       'объект':'№ скв.', 'куст':'куст', 'ГФ':'ГФ',
                       'База дбч_неф':'База нефть',     'ОБД дбч_неф':'ОБД нефть',
                       'База дбч_жидк':'База жидкость', 'ОБД дбч_жидк':'ОБД жидкость'}, inplace=True)
        return pivtab 
       
    def get_pivtab_pane(self, profiles, PNG=False):
        '''
        prof_list - [bazpr, obdpr]
        горизонтальная шапка сводной таблицы с привязкой к объектам
        '''
        pivtab_pane = self.parent.chequer[['ГДМ', 'Номер сектор', 'КУСТ', 'ГФ', 'Группа']].copy()
        pivtab_pane.columns = ['объект', 'сектор', 'куст', 'ГФ', 'группа']
 
        pivtab_pane = pivtab_pane.assign(**{'версия': ''})
 
        if (PNG == True) & (profilefrmt == 'tnav'):
            pivtab_pane = pivtab_pane[['группа', 'версия', 'сектор', 'объект', 'куст', 'ГФ']]
        else:
            pivtab_pane = pivtab_pane[['группа', 'версия', 'сектор', 'объект', 'куст']]
            
        dct_inds = {}

        for profile in profiles:
            dct_inds[profile.name] = list(set(profile['объект']))
        inds = set(sum(dct_inds.values(), [])) & set(self.parent.chequer['ГДМ'])
 
        pivtab_pane = pivtab_pane[pivtab_pane['объект'].isin(inds)]
        
        pivtab_pane.set_index('объект', drop=False, inplace=True)

        return pivtab_pane

    def get_MER_values(self, values):
        '''
        возвращает непустые значения интересующего столбца при пересчете коэф
        values = ИЛИ('деб_неф', 'деб_жидк', 'приемист')
        '''
        MER_values = self.parent.MER[[values]].copy()
        MER_values.dropna(inplace=True)
        return MER_values
    
    
    def write_KRS(self, profiles, pivtab_names, params, sheet_name):
        '''
        parameter = 'КРС'
        '''       
        # горизонтальная шапка сводной таблицы с привязкой к объектам
        pivtab_pane = self.get_pivtab_pane(profiles)
        pivtab_ref.columns = pd.MultiIndex.from_product([[''], [''], pivtab_ref.columns])

        dct_prfls = {}
        
        for profile in profiles:    
            for prm in params:
                dct_prfls[profile.name] = profile.pivot_table(values=prm, index='объект',
                                                              columns=exportfrmt, aggfunc='sum')
    
                # добавляем 2 и 3 уровень шапки
                dct_prfls[profile.name].columns = pd.MultiIndex.from_product\
                    ([[profile.name+' '+prm],[''], dct_prfls[profile.name].columns])
                
            pivtab = pd.concat(dct_prfls.values(), axis=1)
        
        pivtab = pd.merge(left=pivtab_ref, right=pivtab, how='left', 
                      left_index=True, right_index=True, sort=True).fillna(0)
                            
        pivtab.rename(columns = 
                      {'группа':'Группа', 'версия':'Версия профиля', 'сектор':'№ сектора',
                       'объект':'№ скв.', 'куст':'куст', 'ГФ':'ГФ',
                       'База дбч_неф':'База нефть', 'ОБД дбч_неф':'ОБД нефть',
                       'База дбч_жидк':'База жидкость', 'ОБД дбч_жидк':'ОБД жидкость'}, inplace=True)
        
        freeze_panes = (6,4)        
        sheet = pivtab
        sheet.to_excel(self.parent.writer, sheet_name=sheet_name, startrow=3, 
                       index=True, freeze_panes=freeze_panes)
        
    def get_single_pivtab_sheet(self, profile:pd.DataFrame, profile_name:str, prm:str, PNG:bool):
        '''
        profile - bazpr or obdpr
        profile_name 'ОБД' или 'База'
        prm - 'деб_неф', 'деб_жидк', 'приемист'
        '''
        
        # горизонтальная шапка сводной таблицы с привязкой к объектам
        pivtab_pane = self.get_pivtab_pane([profile], PNG)
        
    
        pivtab = profile.pivot_table(values=prm, index='объект',
                                   columns='месяц', aggfunc='sum')
    
        if chbox_dict[profile_name, prm] == 1:
            
            # пересчитываем конечные значения на коэффициенты МЭР,
            # высчитанные с использованием интересующего профиля
            prm_corr = profile[['объект', prm]].set_index('объект')
                    
            prm_corr = prm_corr[prm_corr[prm].ne(0)]\
                .groupby('объект').first() #отсекаем нулевые значения
            
            # подгружаем значения МЭР
            MERval = self.get_MER_values(prm)
            
            # считаем коэффициенты
            coefs = MERval.div(prm_corr).fillna(1)
            coefs.rename(columns={prm:'МЭР_коэф'}, inplace=True)
            
    
            pivtab = pd.merge(how='left', left=pivtab, right=coefs, left_index=True, 
                              right_index=True)
            
            pivtab = pivtab.apply(axis=1, 
                                  func=lambda row: row.iloc[:-1]*row.loc['МЭР_коэф'])

            pivtab = pd.merge(how='left', left=pivtab, right=coefs, left_index=True, 
                              right_index=True)
            
        pivtab = pivtab.fillna(0)
        pivtab = pivtab.loc[~(pivtab.loc[:, pivtab.columns!='МЭР_коэф']==0)\
                            .all(axis=1)]
        #
        # пересечем скважины из получившейся таблицы со скважинами из шахматки    
        pivtab = pd.merge(left=pivtab_pane, right=pivtab, how='inner',
                      left_index=True, right_index=True, sort=True)

       
        # присвоим ИМЯ_OIS для номера скважины
        pivtab['объект'] = self.parent.chequer.set_index('ГДМ').loc[pivtab.index]['ИМЯ_OIS'] 
                       
        pivtab.rename(columns = 
                  {'группа':'Группа', 'версия':'Версия профиля', 'сектор':'№ сектора',
                   'объект':'№ скв.', 'куст':'куст'}, inplace=True)
    
        return pivtab        
#%%
'''
MAIN
'''       
if __name__ == "__main__":
    app = MainApp()
    app.mainloop()